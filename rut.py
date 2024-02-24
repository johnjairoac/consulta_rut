import streamlit as st
import sqlite3
import chromedriver_autoinstaller  # Importar la biblioteca para la instalación automática de ChromeDriver
from selenium import webdriver
from datetime import datetime
import openpyxl
import io
from openpyxl.styles import Border, Side
from selenium.common.exceptions import NoSuchElementException
from openpyxl.utils import get_column_letter

# Instalar automáticamente la versión correcta de ChromeDriver
chromedriver_autoinstaller.install()

# Configurar el título de la pestaña y el ícono
st.set_page_config(page_title=" - Consulta Rut - ", page_icon="consulta.ico")

# Crear la conexión a la base de datos SQLite
conn = sqlite3.connect('user_database.db')
cursor = conn.cursor()

# Crear la tabla de usuarios si no existe
cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY,
        password TEXT
    )
''')
conn.commit()

# Inicializar el estado de sesión
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False

# Función para verificar las credenciales del usuario
def authenticate_user(username, password):
    cursor.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password))
    return cursor.fetchone() is not None

# Página principal con formulario de inicio de sesión
def login_page():
    st.title("Login")

    username = st.text_input("Usuario")
    password = st.text_input("Contraseña", type="password")

    if st.button("Iniciar sesión"):
        if authenticate_user(username, password):
            st.success("Inicio de sesión exitoso!")
            st.session_state.logged_in = True

            st.experimental_rerun()  # Reiniciar la aplicación
        else:
            st.error("Credenciales incorrectas")

# Página de la aplicación después de iniciar sesión
def app_page():
    def main():
        # Configurar opciones para el controlador WebDriver
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')  # Agregar esta línea para ejecutar sin interfaz gráfica
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1200x600')

        # Botón para cerrar sesión con una clave única
        if st.button("Cerrar Sesión", key="cerrar_sesion"):
            st.session_state.logged_in = False
            st.experimental_rerun()  # Reiniciar la aplicación

        # Alinear el botón "Cerrar Sesión" a la derecha
        st.markdown(
            """
            <style>
            div[data-testid="stButton"][data-baseweb="button"] {
                float: right;
            }
            </style>
            """,
            unsafe_allow_html=True
        )

        # Título grande
        st.title("Consulta de Rut")

        # Widget de entrada de texto para los números de NIT
        input_key = "input_nit"
        # Ajusta el tamaño vertical del área de texto con el parámetro height
        nit_input = st.text_area("Ingrese uno o varios números de NIT (uno por línea):", key=input_key, height=200)

        # Contenedor para mostrar el contador de registros
        contador_container = st.empty()

        # Crear un libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Número de NIT", "DV", "Razón Social", "Primer Apellido", "Segundo Apellido", "Primer Nombre",
                      "Otros Nombres", "Estado", "Fecha de Consulta", "Verificación"])

        # Contador de registros ingresados
        registros_ingresados = 0

        # Mostrar en tiempo real el contador de registros y la barra de progreso
        contador_text = contador_container.text(f"Registros ingresados: {registros_ingresados}")
        barra_progreso = st.progress(0)

        # Botón de procesar
        if st.button("Procesar"):
            if not nit_input.strip():  # Verificar si el campo de texto está vacío
                st.warning("No se encuentran NITs ingresados.")
            else:
                # Obtener la cantidad total de registros
                total_registros = len(nit_input.strip().split('\n'))

                # Configuración del controlador WebDriver con las opciones
                driver = webdriver.Chrome(options=chrome_options)

                # Iterar sobre los números de NIT ingresados
                for i, numero_nit in enumerate(nit_input.strip().split('\n'), start=1):
                    # Ignorar campos vacíos
                    if numero_nit.strip() == "":
                        continue

                    registros_ingresados += 1

                    driver.get("https://muisca.dian.gov.co/WebRutMuisca/DefConsultaEstadoRUT.faces")
                    campo_numero = driver.find_element("id", "vistaConsultaEstadoRUT:formConsultaEstadoRUT:numNit")
                    campo_numero.clear()
                    campo_numero.send_keys(numero_nit)
                    boton_buscar = driver.find_element("id", "vistaConsultaEstadoRUT:formConsultaEstadoRUT:btnBuscar")
                    boton_buscar.click()

                    # Definir variables para almacenar la información
                    razon_social = ""
                    estado = ""
                    dv = ""
                    verificacion = ""

                    try:
                        primer_apellido = driver.find_element("id",
                                                              "vistaConsultaEstadoRUT:formConsultaEstadoRUT:primerApellido").text
                    except NoSuchElementException:
                        primer_apellido = ""

                    try:
                        segundo_apellido = driver.find_element("id",
                                                               "vistaConsultaEstadoRUT:formConsultaEstadoRUT:segundoApellido").text
                    except NoSuchElementException:
                        segundo_apellido = ""

                    try:
                        primer_nombre = driver.find_element("id",
                                                            "vistaConsultaEstadoRUT:formConsultaEstadoRUT:primerNombre").text
                    except NoSuchElementException:
                        primer_nombre = ""

                    try:
                        otros_nombres = driver.find_element("id",
                                                            "vistaConsultaEstadoRUT:formConsultaEstadoRUT:otrosNombres").text
                    except NoSuchElementException:
                        otros_nombres = ""

                    if primer_apellido or segundo_apellido or primer_nombre or otros_nombres:
                        razon_social = f"{primer_apellido} {segundo_apellido} {primer_nombre} {otros_nombres}"
                    else:
                        try:
                            razon_social = driver.find_element("id",
                                                               "vistaConsultaEstadoRUT:formConsultaEstadoRUT:razonSocial").text
                        except NoSuchElementException:
                            razon_social = "No existe"

                    try:
                        estado = driver.find_element("id", "vistaConsultaEstadoRUT:formConsultaEstadoRUT:estado").text
                    except NoSuchElementException:
                        estado = ""

                    try:
                        dv = driver.find_element("id", "vistaConsultaEstadoRUT:formConsultaEstadoRUT:dv").text
                    except NoSuchElementException:
                        dv = ""

                    if estado == "REGISTRO ACTIVO":
                        verificacion = "✓"
                    else:
                        verificacion = "X"

                    # Resto del código que procesa la información obtenida y la guarda en el libro de Excel
                    fecha_actual = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
                    sheet.append(
                        [numero_nit, dv, razon_social, primer_apellido, segundo_apellido, primer_nombre, otros_nombres,
                         estado, fecha_actual, verificacion])

                    # Actualizar en tiempo real la cantidad de registros ingresados y la barra de progreso
                    contador_text.text(f"Registros ingresados: {registros_ingresados} de {total_registros}")
                    barra_progreso.progress(i / total_registros)

                # Cierre del controlador WebDriver
                driver.quit()

                # Dar formato al Excel
                for row in sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                fill_type="solid")
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center",
                                                                   wrap_text=True)
                        cell.font = openpyxl.styles.Font(bold=True)

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center",
                                                                   wrap_text=True)

                # Ajustar ancho de columnas
                for column in sheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

                # Agregar bordes a todas las celdas
                border_style = Border(left=Side(border_style="thin", color="000000"),
                                      right=Side(border_style="thin", color="000000"),
                                      top=Side(border_style="thin", color="000000"),
                                      bottom=Side(border_style="thin", color="000000"))

                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = border_style

                # Crear el enlace de descarga
                output = io.BytesIO()
                workbook.save(output)
                output.seek(0)
                st.download_button(label="Descargar Excel", data=output, key="download_excel",
                                   file_name=f"consulta_rut_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        # Botón de cancelar
        if st.button("Cancelar"):
            st.write("Operación cancelada. No se realizaron búsquedas.")

    if __name__ == "__main__":
        main()

# Lógica para determinar la página a mostrar
if st.session_state.logged_in:
    app_page()
else:
    login_page()
